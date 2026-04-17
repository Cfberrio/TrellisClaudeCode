import csv
import json
from pathlib import Path

import pytest

from src.common.io import output_dir, write_csv, write_json, write_md, write_xlsx


def test_output_dir_creates_path(tmp_path):
    d = output_dir(tmp_path, "discipline_rift", "2026-04")
    assert d == tmp_path / "output" / "meta" / "discipline_rift" / "2026-04"
    assert d.exists()


def test_output_dir_idempotent(tmp_path):
    d1 = output_dir(tmp_path, "discipline_rift", "2026-04")
    d2 = output_dir(tmp_path, "discipline_rift", "2026-04")
    assert d1 == d2


def test_write_json(tmp_path):
    data = {"brand": "test", "value": 42}
    path = tmp_path / "test.json"
    write_json(path, data)
    assert json.loads(path.read_text()) == data


def test_write_json_handles_none_values(tmp_path):
    data = {"brand": "test", "value": None}
    path = tmp_path / "test.json"
    write_json(path, data)
    assert json.loads(path.read_text())["value"] is None


def test_write_csv(tmp_path):
    rows = [
        {"brand": "dr", "metric": "reach", "value": "100"},
        {"brand": "dr", "metric": "likes", "value": "50"},
    ]
    path = tmp_path / "test.csv"
    write_csv(path, rows, fieldnames=["brand", "metric", "value"])
    content = path.read_text()
    assert "brand,metric,value" in content
    assert "dr,reach,100" in content


def test_write_md(tmp_path):
    path = tmp_path / "report.md"
    write_md(path, "# Title\n\nBody text.")
    assert path.read_text() == "# Title\n\nBody text."


def test_write_xlsx_creates_file(tmp_path):
    path = tmp_path / "report.xlsx"
    sheets = {
        "Overview": [{"metric": "Followers", "value": 1234}],
        "Empty Sheet": [],
    }
    write_xlsx(path, sheets)
    assert path.exists()
    assert path.stat().st_size > 0


def test_write_xlsx_correct_sheet_names(tmp_path):
    import openpyxl
    path = tmp_path / "report.xlsx"
    write_xlsx(path, {"Facebook": [{"a": 1}], "Instagram": [{"b": 2}]})
    wb = openpyxl.load_workbook(path)
    assert "Facebook" in wb.sheetnames
    assert "Instagram" in wb.sheetnames
