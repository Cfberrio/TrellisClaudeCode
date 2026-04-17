"""
io.py — File I/O utilities for the Meta monthly pipeline.

Handles writing JSON, CSV, xlsx, and markdown outputs.
All paths are resolved relative to a provided base directory.
"""

import csv
import json
from pathlib import Path
from typing import Any

# Canonical field order for monthly_metrics_flat.csv
FLAT_FIELDNAMES = [
    "brand",
    "month",
    "platform",
    "source_type",
    "entity_id",
    "entity_name",
    "media_id",
    "media_type",
    "media_product_type",
    "metric_name",
    "metric_value",
    "date",
    "permalink",
    "timestamp",
    "caption_snippet",
]


def output_dir(base: Path, brand_slug: str, month_str: str) -> Path:
    """
    Resolve and create output/meta/{brand_slug}/{month_str}/ relative to base.

    Returns the created Path.
    """
    d = base / "output" / "meta" / brand_slug / month_str
    d.mkdir(parents=True, exist_ok=True)
    return d


def write_json(path: Path, data: Any) -> None:
    """Write data as indented UTF-8 JSON."""
    path.write_text(
        json.dumps(data, indent=2, ensure_ascii=False, default=str),
        encoding="utf-8",
    )


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    """Write rows as CSV with the given field order. Extra keys are ignored."""
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_md(path: Path, content: str) -> None:
    """Write a string to a markdown file."""
    path.write_text(content, encoding="utf-8")


def write_xlsx(path: Path, sheets: dict[str, list[dict]]) -> None:
    """
    Write an xlsx workbook with one sheet per key in sheets.

    Each sheet value is a list of dicts. Dict keys become column headers.
    Empty lists produce a sheet with a 'No data available' placeholder.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise ImportError(
            "openpyxl is required for xlsx output. "
            "Install with: pip install openpyxl"
        )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    header_fill = PatternFill("solid", fgColor="E8F0FE")
    header_font = Font(bold=True)

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel tab name limit

        if not rows:
            ws.append(["No data available"])
            continue

        headers = list(rows[0].keys())
        ws.append(headers)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        for row in rows:
            ws.append([row.get(h) for h in headers])

        # Auto-fit column widths (capped at 50)
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value or "")) for cell in col), default=8
            )
            ws.column_dimensions[col[0].column_letter].width = min(
                max_len + 2, 50
            )

    wb.save(path)
